import openpyxl
from xlrd import sheet


class Util:
    @classmethod
    def get_test_data_from_txt(cls,path):
        cmds=[]
        with open(path,'r',encoding='utf8') as f:
            while True:
                line=f.readline()
                if not line:
                    break
                if not line.startswith('#'):
                    cmds.append(line.strip())  # 加入列表，去掉前后空格
            return cmds

    @classmethod
    def get_test_data_from_excel(cls, path):
        book=openpyxl.load_workbook(path)
        sheet=book.active
        test_data=[]
        for row in range(2,sheet.max_row+1):
            cmd = ''
            for col in range(2, sheet.max_row + 1):
                content=sheet.cell(row=row,column=col).value
                if content:
                    cmd+=str(content)
                    cmd+='#'
            test_data.append(cmd[:-1])
        print(test_data)
